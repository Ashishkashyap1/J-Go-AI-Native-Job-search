"""
Excel Export - Exports job search results to a formatted Excel file.
"""

import os
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("litsearch")


# Color scheme
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
EXCELLENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MODERATE_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
POOR_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
LINK_FONT = Font(name="Calibri", color="0563C1", underline="single", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_score_fill(score: float) -> PatternFill:
    """Return background fill based on score."""
    if score >= 80:
        return EXCELLENT_FILL
    elif score >= 60:
        return GOOD_FILL
    elif score >= 40:
        return MODERATE_FILL
    elif score >= 20:
        return LOW_FILL
    return POOR_FILL


def export_to_excel(jobs: list[dict], cv_data: dict, output_path: str = None,
                    portal_counts: dict = None, qc_report: dict = None) -> str:
    """
    Export job results to a formatted Excel file.

    Args:
        jobs: List of job dicts with scores
        cv_data: CV parsed data
        output_path: Output file path (auto-generated if None)
        portal_counts: Dict of portal -> job count

    Returns:
        str: Path to the created Excel file
    """
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"LITSEARCH_Results_{timestamp}.xlsx"

    wb = Workbook()

    # --- Sheet 1: Top Matches ---
    ws_top = wb.active
    ws_top.title = "Top Matches"

    # Header row
    headers = [
        "Rank", "Company Name", "Job Role", "Location", "Source",
        "Score Basis", "Score / 100", "Match Level",
        "Missing Skills", "QC Confidence", "Link to Application"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_top.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws_top.row_dimensions[1].height = 30

    # Data rows - sorted by score
    sorted_jobs = sorted(jobs, key=lambda j: j.get("score", 0), reverse=True)

    for idx, job in enumerate(sorted_jobs, 1):
        row = idx + 1
        score = job.get("score", 0)

        missing = ", ".join(job.get("missing_skills", [])[:5]) or "—"

        values = [
            idx,
            job.get("company", "N/A"),
            job.get("title", "N/A"),
            job.get("location", "N/A"),
            job.get("source", "N/A"),
            job.get("score_basis", "n/a"),
            score,
            _get_match_label(score),
            missing,
            job.get("qc_confidence", "n/a"),
            job.get("url", ""),
        ]

        for col, value in enumerate(values, 1):
            cell = ws_top.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Color the score column
            if col == 7:  # Score column
                cell.fill = get_score_fill(score)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 8:  # Match level
                cell.fill = get_score_fill(score)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 6:  # Match percentage
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 11 and value:  # Link column
                cell.font = LINK_FONT
                cell.hyperlink = value
                cell.value = "Apply Here"

    # Set column widths
    widths = [6, 30, 35, 20, 12, 20, 12, 18, 25, 15, 15]
    for col, width in enumerate(widths, 1):
        ws_top.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws_top.freeze_panes = "A2"

    # --- Sheet 2: All Jobs ---
    ws_all = wb.create_sheet("All Jobs")

    all_headers = [
        "S.No", "Company Name", "Job Role", "Description (snippet)",
        "Location", "Source Portal", "Matched Skills", "Missing Skills",
        "Score", "QC Flags", "Job Link"
    ]

    for col, header in enumerate(all_headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws_all.row_dimensions[1].height = 30

    for idx, job in enumerate(sorted_jobs, 1):
        row = idx + 1
        score = job.get("score", 0)

        values = [
            idx,
            job.get("company", "N/A"),
            job.get("title", "N/A"),
            job.get("description", "N/A")[:200] if job.get("description") else "N/A",
            job.get("location", "N/A"),
            job.get("source", "N/A"),
            ", ".join(job.get("matched_skills", [])) or "—",
            ", ".join(job.get("missing_skills", [])) or "—",
            score,
            ", ".join(job.get("qc_flags", [])) or "clean",
            job.get("url", ""),
        ]

        for col, value in enumerate(values, 1):
            cell = ws_all.cell(row=row, column=col, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 9:
                cell.fill = get_score_fill(score)
            elif col == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 11 and value:
                cell.font = LINK_FONT
                cell.hyperlink = value
                cell.value = "Apply Here"

    all_widths = [6, 30, 35, 50, 20, 14, 10, 20, 10, 15, 15]
    for col, width in enumerate(all_widths, 1):
        ws_all.column_dimensions[get_column_letter(col)].width = width
    ws_all.freeze_panes = "A2"

    # --- Sheet 3: CV Profile Summary ---
    ws_cv = wb.create_sheet("CV Profile")

    ws_cv.cell(row=1, column=1, value="LITSEARCH - CV Profile Summary").font = Font(
        name="Calibri", bold=True, size=14, color="1F4E79"
    )
    ws_cv.merge_cells("A1:D1")

    profile_data = [
        ("Name", cv_data.get("name", "N/A")),
        ("Email", cv_data.get("email", "N/A")),
        ("Location", cv_data.get("location", "N/A")),
        ("", ""),
        ("Extracted Skills", ", ".join(cv_data.get("skills", []))),
        ("Suggested Roles", ", ".join(cv_data.get("suggested_roles", []))),
        ("", ""),
        ("Experience", ""),
    ]

    for exp in cv_data.get("experience", []):
        profile_data.append((
            f"  {exp.get('company', '')}",
            f"{exp.get('role', '')} ({exp.get('duration', '')})"
        ))

    profile_data.append(("", ""))
    profile_data.append(("Education", ", ".join(cv_data.get("education", []))))
    profile_data.append(("Certifications", ", ".join(cv_data.get("certifications", []))))

    for row_idx, (label, value) in enumerate(profile_data, 3):
        ws_cv.cell(row=row_idx, column=1, value=label).font = BOLD_FONT
        ws_cv.cell(row=row_idx, column=2, value=value).font = NORMAL_FONT

    ws_cv.column_dimensions["A"].width = 25
    ws_cv.column_dimensions["B"].width = 80

    # --- Sheet 4: Portal Summary ---
    if portal_counts:
        ws_portal = wb.create_sheet("Portal Summary")
        ws_portal.cell(row=1, column=1, value="Jobs Found Per Portal").font = Font(
            name="Calibri", bold=True, size=14, color="1F4E79"
        )
        ws_portal.merge_cells("A1:C1")

        ws_portal.cell(row=3, column=1, value="Portal").font = HEADER_FONT
        ws_portal.cell(row=3, column=1).fill = HEADER_FILL
        ws_portal.cell(row=3, column=2, value="Jobs Found").font = HEADER_FONT
        ws_portal.cell(row=3, column=2).fill = HEADER_FILL

        for row_idx, (portal, count) in enumerate(sorted(portal_counts.items()), 4):
            ws_portal.cell(row=row_idx, column=1, value=portal.title()).font = NORMAL_FONT
            ws_portal.cell(row=row_idx, column=2, value=count).font = NORMAL_FONT

        total_row = 4 + len(portal_counts)
        ws_portal.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
        ws_portal.cell(row=total_row, column=2, value=sum(portal_counts.values())).font = BOLD_FONT

        ws_portal.column_dimensions["A"].width = 20
        ws_portal.column_dimensions["B"].width = 15

    # --- Sheet 5: QC Report (maker-checker audit trail) ---
    if qc_report:
        ws_qc = wb.create_sheet("QC Report")
        ws_qc.cell(row=1, column=1, value="Maker-Checker QC Report").font = Font(
            name="Calibri", bold=True, size=14, color="1F4E79")
        ws_qc.merge_cells("A1:E1")

        r = 3
        ws_qc.cell(row=r, column=1, value="Scraped records").font = BOLD_FONT
        ws_qc.cell(row=r, column=2, value=len(qc_report.get("accepted", []))
                   + len(qc_report.get("rejected", [])))
        r += 1
        ws_qc.cell(row=r, column=1, value="Accepted").font = BOLD_FONT
        ws_qc.cell(row=r, column=2, value=len(qc_report.get("accepted", [])))
        r += 1
        ws_qc.cell(row=r, column=1, value="Rejected").font = BOLD_FONT
        ws_qc.cell(row=r, column=2, value=len(qc_report.get("rejected", [])))
        r += 2

        if qc_report.get("dead_portals"):
            cell = ws_qc.cell(row=r, column=1,
                value="DEAD PORTALS (0 results — scraper broken/blocked): "
                      + ", ".join(qc_report["dead_portals"]))
            cell.font = Font(name="Calibri", bold=True, color="9C0006", size=11)
            cell.fill = POOR_FILL
            ws_qc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 2

        if qc_report.get("reject_reasons"):
            ws_qc.cell(row=r, column=1, value="Reject reasons").font = BOLD_FONT
            r += 1
            for reason, cnt in sorted(qc_report["reject_reasons"].items()):
                ws_qc.cell(row=r, column=1, value=reason).font = NORMAL_FONT
                ws_qc.cell(row=r, column=2, value=cnt).font = NORMAL_FONT
                r += 1
            r += 1

        if qc_report.get("rejected"):
            headers_qc = ["Rejected Title", "Company", "Source", "Reasons", "URL"]
            for c, h in enumerate(headers_qc, 1):
                cell = ws_qc.cell(row=r, column=c, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            r += 1
            for rj in qc_report["rejected"]:
                ws_qc.cell(row=r, column=1, value=rj.get("title", "")).font = NORMAL_FONT
                ws_qc.cell(row=r, column=2, value=rj.get("company", "")).font = NORMAL_FONT
                ws_qc.cell(row=r, column=3, value=rj.get("source", "")).font = NORMAL_FONT
                ws_qc.cell(row=r, column=4,
                           value=", ".join(rj.get("qc_reasons", []))).font = NORMAL_FONT
                ws_qc.cell(row=r, column=5, value=rj.get("url", "")).font = NORMAL_FONT
                r += 1

        for col, width in zip("ABCDE", [45, 25, 14, 30, 60]):
            ws_qc.column_dimensions[col].width = width

    # Save
    wb.save(output_path)
    return output_path


def _get_match_label(score: float) -> str:
    """Return a human-readable label for a match score."""
    if score >= 80:
        return "🟢 Excellent"
    elif score >= 60:
        return "🔵 Good"
    elif score >= 40:
        return "🟡 Moderate"
    elif score >= 20:
        return "🟠 Low"
    return "🔴 Poor"
